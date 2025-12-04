from django.shortcuts import render
from .serializer import agendarClaseSerializer, datosFisicosSerializer, registroSerializer, coachSerializer, membresiaSerializer, nuevoCoachSerializer, claseProgramadaSerializer, pagoSerializer
from .models import AgendarClase, DatosFisicos, Registro, Coach, Membresia, NuevoCoach, ClaseProgramada, Pago
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from django.utils import timezone
from rest_framework.parsers import MultiPartParser, FormParser
import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse

# Create your views here.
@api_view(['GET', 'POST'])
def agendarClase_list(request):
    if request.method == 'GET':
        clases = AgendarClase.objects.all()
        serializer = agendarClaseSerializer(clases, many=True)
        return Response(serializer.data)
    
    if request.method == 'POST':
        serializer = agendarClaseSerializer(data = request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    

@api_view(['GET', 'POST'])
def datosFisicos_list(request):
    if request.method == 'GET':
        datos = DatosFisicos.objects.all()
        serializer = datosFisicosSerializer(datos, many=True)
        return Response(serializer.data)
    
    if request.method == 'POST':
        serializer = datosFisicosSerializer(data = request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
def datosFisicos_usuario(request, user_id):
    datos = DatosFisicos.objects.filter(usuario_id=user_id).order_by('-fecha')
    serializer = datosFisicosSerializer(datos, many=True)
    return Response(serializer.data)


@api_view(['GET', 'POST', 'PATCH'])
def registro_list(request):
    if request.method == 'GET':
        registros = Registro.objects.all()
        serializer = registroSerializer(registros, many=True)
        return Response(serializer.data)
    
    if request.method == 'POST':
        serializer = registroSerializer(data = request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    if request.method == 'PATCH':
        user_id = request.data.get("id")
        if not user_id:
            return Response({"error": "Debe enviar id del usuario"}, status=status.HTTP_400_BAD_REQUEST)
        try:
            registro = Registro.objects.get(id=user_id)
        except Registro.DoesNotExist:
            return Response({"error": "Usuario no encontrado"}, status=status.HTTP_404_NOT_FOUND)
        serializer = registroSerializer(registro, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
@api_view(['GET', 'POST', 'PATCH'])
def coach_list(request):
    if request.method == 'GET':
        entrenador = Coach.objects.all()
        serializer = coachSerializer(entrenador, many=True)
        return Response(serializer.data)
    
    if request.method == 'POST':
        serializer = coachSerializer(data = request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    if request.method == 'PATCH':
        serializer = coachSerializer(Coach, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=400)
    
    
@api_view(['POST'])
def membresia_list(request):
    usuario_id = request.data.get("usuario")

    usuario = Registro.objects.get(id=usuario_id)

    activa = Membresia.objects.filter(usuario=usuario, is_active=True).first()
    if activa:
        return Response({"error": "Ya tienes una membresía activa"}, status=400)

    serializer = membresiaSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=201)
    return Response(serializer.errors, status=400)

@api_view(['POST'])
def comprar_membresia_list(request):
    # 1. Recibir datos (incluyendo archivos)
    usuario_id = request.data.get("usuario")

    try:
        usuario_instancia = Registro.objects.get(id=usuario_id)
    except Registro.DoesNotExist:
        return Response({"error": "Usuario no encontrado"}, status=404)

    # 2. Crear membresía en estado Pendiente
    nueva_membresia = Membresia.objects.create(
        usuario=usuario_instancia,
        plan_nombre=request.data.get("plan_nombre"),
        plan_clases=request.data.get("plan_clases"),
        plan_precio=request.data.get("plan_precio"),
        comprobante=request.FILES.get('comprobante'),
        estado='Pendiente',
        is_active=False
    )

    # 3. 🔥 Crear registro de pago automático
    Pago.objects.create(
        usuario=usuario_instancia,
        tipo="membresía",
        monto=request.data.get("plan_precio"),
        metodo="transferencia",   # Puedes cambiarlo si lo deseas
        estado="Pendiente"
    )

    return Response({"mensaje": "Solicitud enviada"}, status=201)


@api_view(['GET'])
def membresia_activa(request, user_id):
    try:
        usuario = Registro.objects.get(id=user_id)
    except Registro.DoesNotExist:
        return Response({"activa": False})

    membresia = Membresia.objects.filter(usuario=usuario, is_active=True).first()

    if not membresia:
        return Response({"activa": False})

    return Response({
        "activa": True,
        "plan_nombre": membresia.plan_nombre,
        "plan_clases": membresia.plan_clases,
        "plan_precio": membresia.plan_precio,
        "start_date": membresia.start_date,
        "end_date": membresia.end_date,
    })

# NUEVA VISTA: Para que el Coach apruebe
@api_view(['POST'])
def aprobar_membresia(request, membresia_id):
    try:
        membresia = Membresia.objects.get(id=membresia_id)
    except Membresia.DoesNotExist:
        return Response({"error": "Membresía no encontrada"}, status=404)

    # 1. Activar la membresía
    membresia.estado = 'Activa'
    membresia.is_active = True
    membresia.save()

    # 2. Buscar el pago pendiente asociado a esta membresía
    pago = Pago.objects.filter(
        usuario=membresia.usuario,
        tipo="membresía",
        estado="Pendiente"
    ).first()

    # 3. Actualizar el pago si existe
    if pago:
        pago.estado = "Realizado"
        pago.save()

    return Response({"mensaje": "Membresía aprobada y pago confirmado"})


# NUEVA VISTA: Listar todas las membresías (para el panel del coach)
@api_view(['GET'])
def listar_todas_membresias(request):
    membresias = Membresia.objects.all().order_by('-start_date')
    data = []
    for m in membresias:
        data.append({
            "id": m.id,
            "user": f"{m.usuario.nombre} {m.usuario.apellido}",
            "plan_nombre": m.plan_nombre,
            "plan_precio": m.plan_precio,
            "plan_clases": m.plan_clases,
            "start_date": m.start_date,
            "status": m.estado,
            "comprobante": m.comprobante.url if m.comprobante else None
        })
    return Response(data)


@api_view(['POST'])
def login_usuario(request):
    correo = request.data.get("correo")
    password = request.data.get("password")

    try:
        usuario = Registro.objects.get(correo=correo, password=password)
    except Registro.DoesNotExist:
        return Response({"error": "Credenciales incorrectas"}, status=400)

    return Response({
        "id": usuario.id,
        "nombre": usuario.nombre,
        "correo": usuario.correo,
        'apellido': usuario.apellido
    })

@api_view(['DELETE'])
def eliminar_membresia(request, membresia_id):
    try:
        membresia = Membresia.objects.get(id=membresia_id)
        membresia.delete()
        return Response({"mensaje": "Membresía eliminada"}, status=status.HTTP_204_NO_CONTENT)
    except Membresia.DoesNotExist:
        return Response({"error": "Membresía no encontrada"}, status=status.HTTP_404_NOT_FOUND)

@api_view(['GET', 'POST'])
def coach_list(request):
    if request.method == 'GET':
        coaches = NuevoCoach.objects.all()
        serializer = nuevoCoachSerializer(coaches, many=True)
        return Response(serializer.data)

    if request.method == 'POST':
        serializer = nuevoCoachSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=201)
        return Response(serializer.errors, status=400)
    
@api_view(['GET', 'PUT', 'DELETE'])
def coach_detail(request, coach_id):

    try:
        coach = NuevoCoach.objects.get(id=coach_id)
    except NuevoCoach.DoesNotExist:
        return Response({"error": "Coach no encontrado"}, status=404)

    if request.method == 'GET':
        serializer = nuevoCoachSerializer(coach)
        return Response(serializer.data)

    if request.method == 'PUT':
        serializer = nuevoCoachSerializer(coach, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=400)

    if request.method == 'DELETE':
        coach.delete()
        return Response({"mensaje": "Coach eliminado"}, status=204)
    

@api_view(['GET', 'POST'])
def horario_list(request):
    if request.method == 'GET':
        clases = ClaseProgramada.objects.all().order_by('dia', 'horario')
        serializer = claseProgramadaSerializer(clases, many=True)
        return Response(serializer.data)

    if request.method == 'POST':
        serializer = claseProgramadaSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=201)
        return Response(serializer.errors, status=400)

@api_view(['DELETE'])
def horario_delete(request, pk):
    try:
        clase = ClaseProgramada.objects.get(id=pk)
    except ClaseProgramada.DoesNotExist:
        return Response({"error": "Clase no encontrada"}, status=404)

    clase.delete()
    return Response({"mensaje": "Clase eliminada"}, status=204)


@api_view(['GET'])
def pagos_list(request):
    fecha_inicio = request.GET.get("inicio")
    fecha_fin = request.GET.get("fin")

    # Mostrar SOLO pagos realizados
    pagos = Pago.objects.filter(estado="Realizado").order_by("-fecha", "-hora")

    # Filtro por fecha
    if fecha_inicio and fecha_fin:
        pagos = pagos.filter(fecha__range=[fecha_inicio, fecha_fin])

    serializer = pagoSerializer(pagos, many=True)
    return Response(serializer.data)


@api_view(['GET'])
def exportar_pagos(request):
    fecha_inicio = request.GET.get("inicio")
    fecha_fin = request.GET.get("fin")

    # Mostrar solo pagos realizados
    pagos = Pago.objects.filter(estado="Realizado")

    if fecha_inicio and fecha_fin:
        pagos = pagos.filter(fecha__range=[fecha_inicio, fecha_fin])

    # Crear Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pagos"

    # Encabezados
    headers = ["Usuario", "Tipo", "Monto", "Método", "Fecha", "Hora", "Estado"]
    ws.append(headers)

    # Encabezados en negrita
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Datos
    for p in pagos:
        ws.append([
            f"{p.usuario.nombre} {p.usuario.apellido}",  # Nombre completo
            p.tipo,
            p.monto,
            p.metodo,
            p.fecha.strftime("%d-%m-%Y"),
            p.hora.strftime("%H:%M"),
            p.estado
        ])

    from openpyxl.utils import get_column_letter

    for column_cells in ws.columns:
        max_length = 0

        for cell in column_cells:
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass

        adjusted_width = max_length + 2
        column_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[column_letter].width = adjusted_width

    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    response["Content-Disposition"] = 'attachment; filename="pagos.xlsx"'
    return response

